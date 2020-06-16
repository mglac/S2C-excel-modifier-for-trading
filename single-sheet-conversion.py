# single-sheet-conversion.py - Test single sheet compression
# to prepare an excel sheet for the s2c stock trading platform

import os
from openpyxl import styles
from openpyxl import load_workbook
from openpyxl import Workbook
import tkinter
from tkinter import filedialog
from tkinter import Text
# Open the source workbook from S2C
# Open a new workbook for trade compression


def excel_modifications(filepath):
    initial_wb = load_workbook(filepath, data_only=True, read_only=True)
    for ws in initial_wb:
        if ws.title == "CORE":
            core_compression(ws)
        if ws.title == "Infl":
            infl_compression(ws)
        if ws.title != "AA" and ws.title != "CORE" and ws.title != "HYE" and ws.title != "Infl" and ws.title != "Int" and ws.title != "HYE" and ws.title != "Alts" and ws.title != "ESG" and ws.title != "ML_LTG" and ws.title != "529":
            common_compression(ws)


def common_compression(worksheet):
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = worksheet.title + ".xlsx"
    new_ws_securities_list = []
    new_ws_percents_list = []
    for row in worksheet.values:
        current_percent = row[0]
        current_security = row[1]
        if isinstance(current_percent, float) and isinstance(current_security, str):
            new_ws_securities_list.append(str(current_security))
            new_ws_percents_list.append(current_percent)

    for i in range(len(new_ws_securities_list) + 1):
        if i == 0:
            new_ws.cell(1, 1).value = 'Security'
            new_ws.cell(1, 2).value = '%'
        else:
            new_ws.cell(i+1, 1).value = new_ws_securities_list[i-1]
            new_ws.cell(i+1, 2).value = new_ws_percents_list[i-1]
            new_ws.cell(i+1, 1).number_format = '0.000%'
            new_ws.cell(i+1, 2).number_format = '0.000%'
        new_wb.save(str(new_ws.title))


def core_compression(worksheet):
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = worksheet.title + ".xlsx"
    new_ws_securities_list = []
    new_ws_percents_list = []
    for row in worksheet.values:
        current_percent = row[3]
        current_security = row[0]
        if isinstance(current_percent, float) and isinstance(current_security, str):
            new_ws_securities_list.append(str(current_security))
            new_ws_percents_list.append(current_percent)

    for i in range(len(new_ws_securities_list) + 1):
        if i == 0:
            new_ws.cell(1, 1).value = 'Security'
            new_ws.cell(1, 2).value = '%'
        else:
            new_ws.cell(i+1, 1).value = new_ws_securities_list[i-1]
            new_ws.cell(i+1, 2).value = new_ws_percents_list[i-1]
            new_ws.cell(i+1, 1).number_format = '0.000%'
            new_ws.cell(i+1, 2).number_format = '0.000%'
        new_wb.save(str(new_ws.title))


def infl_compression(worksheet):
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = worksheet.title + ".xlsx"
    new_ws_securities_list = []
    new_ws_percents_list = []
    for row in worksheet.values:
        current_percent = row[2]
        current_security = row[0]
        if isinstance(current_percent, float) and isinstance(current_security, str):
            new_ws_securities_list.append(str(current_security))
            new_ws_percents_list.append(current_percent)

    for i in range(len(new_ws_securities_list) + 1):
        if i == 0:
            new_ws.cell(1, 1).value = 'Security'
            new_ws.cell(1, 2).value = '%'
        else:
            new_ws.cell(i+1, 1).value = new_ws_securities_list[i-1]
            new_ws.cell(i+1, 2).value = new_ws_percents_list[i-1]
            new_ws.cell(i+1, 1).number_format = '0.000%'
            new_ws.cell(i+1, 2).number_format = '0.000%'
        new_wb.save(str(new_ws.title))


def open_file_path():
    filename = filedialog.askopenfilename(
        initialdir="/", title="Select File", filetypes=(("excel workbooks", "*.xlsx"), ("all files", ".")))
    excel_modifications(filename)


window = tkinter.Tk()
window.geometry("300x250")
window.resizable(0, 0)
window.title("Test Window")
tkinter.Label(text="").pack()
open_file = tkinter.Button(text="Select a workbook",
                           height="2", width="30", command=open_file_path)
open_file.pack()
tkinter.Label(window, text="Your new worksheet will be generated off of the"
              + "\nworkbook you select by pressing the button"
              + "\nabove. This will only generate a new workbook"
              + "\nand sheet based off of worksheet 1E in your"
              + "\nselected workbook. The new workbook will generate in"
              + "\nthe same folder the inputted workbook resides in.").pack()


window.mainloop()
