# modifiers.py - This class contains the methods used to alter the portfolio
# spreadsheets to the desired format for trading assets.

import os
from openpyxl import styles
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import filedialog
from tkinter import Text

# excel_modifications Method
# This method takes the file path to the workbook being read along with the
# file path to the location the user wants the new workbooks to be stored at.


def excel_modifications(file_path, storage_path):
    init_wb = load_workbook(file_path, data_only=True)  # loads workbook
    for ws in init_wb:
        if ws.title == "CORE":
            # checks if the current worksheet is the CORE worksheet
            core_compression(ws, storage_path)
        if ws.title == "Infl":
            # checks if the current worksheet is the Infl worksheet
            infl_compression(ws, storage_path)
        if ws.title == "FI" or ws.sheet_properties.tabColor.theme == 9:
            # checks if the current worksheet is the FI worksheet or if the
            # current worksheet's tab number is colored green
            common_compression(ws, storage_path)

# common_compression method
# This method is used to modify any portfolio worksheet along with the Fixed
# Income worksheet. This is done to prepare the worksheet for trading.


def common_compression(worksheet, storage_path):
    new_wb = Workbook()  # creates a new workbook to write to
    new_ws = new_wb.active  # opens to the active worksheet in the new workbook
    new_ws.title = worksheet.title + ".xlsx"  # creates the new workbook title
    new_ws_ticker_list = []  # a list of tickers
    new_ws_percent_list = []  # a list of percents corresponding to tickers
    for row in worksheet.values:  # Iterates through all of the rows in the ws
        curr_percent = row[0]
        curr_ticker = row[1]
        if isinstance(curr_percent, float) and isinstance(curr_ticker, str):
            new_ws_ticker_list.append(str(curr_ticker))
            new_ws_percent_list.append(curr_percent)

    for i in range(len(new_ws_ticker_list) + 1):
        if i == 0:
            # Titles the column as Security
            new_ws.cell(1, 1).value = 'Security'
            new_ws.cell(1, 2).value = '%'  # Titles the column as %
        else:
            new_ws.cell(i+1, 1).value = new_ws_ticker_list[i-1]
            new_ws.cell(i+1, 2).value = new_ws_percent_list[i-1]
            new_ws.cell(i+1, 1).number_format = '0.000%'
            new_ws.cell(i+1, 2).number_format = '0.000%'
        # Saves in specified file location
        new_wb.save(storage_path + "\\" + str(new_ws.title))

# core_compression method
# This method is used to modify the Core Stocks worksheet. This is done to
# prepare the worksheet for trading.


def core_compression(worksheet, storage_path):
    new_wb = Workbook()  # creates a new workbook to write to
    new_ws = new_wb.active  # opens to the active worksheet in the new workbook
    new_ws.title = worksheet.title + ".xlsx"  # creates the new workbook title
    new_ws_ticker_list = []  # a list of tickers
    new_ws_percent_list = []  # a list of percents corresponding to tickers
    for row in worksheet.values:  # Iterates through all of the rows in the ws
        curr_percent = row[3]
        curr_ticker = row[0]
        if isinstance(curr_percent, float) and isinstance(curr_ticker, str):
            new_ws_ticker_list.append(str(curr_ticker))
            new_ws_percent_list.append(curr_percent)

    for i in range(len(new_ws_ticker_list) + 1):
        if i == 0:
            # Titles the column as Security
            new_ws.cell(1, 1).value = 'Security'
            new_ws.cell(1, 2).value = '%'  # Titles the column as %
        else:
            new_ws.cell(i+1, 1).value = new_ws_ticker_list[i-1]
            new_ws.cell(i+1, 2).value = new_ws_percent_list[i-1]
            new_ws.cell(i+1, 1).number_format = '0.000%'
            new_ws.cell(i+1, 2).number_format = '0.000%'
        # Saves in specified file location
        new_wb.save(storage_path + "\\" + str(new_ws.title))

# infl_compression method
# This method is used to modify the Inflation Model worksheet. This is done to
# prepare the worksheet for trading.


def infl_compression(worksheet, storage_path):
    new_wb = Workbook()  # creates a new workbook to write to
    new_ws = new_wb.active  # opens to the active worksheet in the new workbook
    new_ws.title = worksheet.title + ".xlsx"  # creates the new workbook title
    new_ws_ticker_list = []  # a list of tickers
    new_ws_percent_list = []  # a list of percents corresponding to tickers
    for row in worksheet.values:  # Iterates through all of the rows in the ws
        curr_percent = row[2]
        curr_ticker = row[0]
        if isinstance(curr_percent, float) and isinstance(curr_ticker, str):
            new_ws_ticker_list.append(str(curr_ticker))
            new_ws_percent_list.append(curr_percent)

    for i in range(len(new_ws_ticker_list) + 1):
        if i == 0:
            # Titles the column as Security
            new_ws.cell(1, 1).value = 'Security'
            new_ws.cell(1, 2).value = '%'  # Titles the column as %
        else:
            new_ws.cell(i+1, 1).value = new_ws_ticker_list[i-1]
            new_ws.cell(i+1, 2).value = new_ws_percent_list[i-1]
            new_ws.cell(i+1, 1).number_format = '0.000%'
            new_ws.cell(i+1, 2).number_format = '0.000%'
        # Saves in specified file location
        new_wb.save(storage_path + "\\" + str(new_ws.title))
